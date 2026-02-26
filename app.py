"""
Client Balance - Sistema de controle financeiro por cliente
Banco de dados: SQLite (banco.db)
Interface: CustomTkinter
"""

import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, Listbox as tk_Listbox
from datetime import date
import customtkinter as ctk
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ─── Tema padrão ────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


# ═══════════════════════════════════════════════════════════════════════════
# UTILITARIOS
# ═══════════════════════════════════════════════════════════════════════════

def formatar_moeda(valor_float: float) -> str:
    """Converte float para string no formato R$ 1.234,56."""
    return f"R$ {valor_float:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def aplicar_mascara_moeda(entry: ctk.CTkEntry, var: ctk.StringVar = None):
    """
    Aplica máscara de moeda BR no entry via KeyRelease.
    - O usuário digita apenas números (vírgula/ponto são ignorados).
    - Os dígitos acumulados são exibidos como inteiro com separador de milhar.
    - Ex: digitar 1 -> "1", digitar 500 -> "1.500", digitar 00 -> "1.500,00"
      (sem modo centavos — o usuário digita o valor exato inteiro ou com vírgula manual)

    Estratégia simples e sem bug de cursor:
    - Mantém apenas dígitos e UMA vírgula opcional.
    - Formata a parte inteira com pontos de milhar.
    - Cursor sempre vai ao final após formatação.
    """
    _bloqueado = [False]

    def _formatar(event=None):
        if _bloqueado[0]:
            return
        _bloqueado[0] = True
        try:
            raw = entry.get()

            # Separa parte inteira e decimal (aceita vírgula ou ponto como separador)
            raw_norm = raw.replace(".", "").replace(",", ".")
            if "." in raw_norm:
                partes = raw_norm.split(".")
                inteiros = "".join(c for c in partes[0] if c.isdigit())
                decimais = "".join(c for c in partes[1] if c.isdigit())[:2]
            else:
                inteiros = "".join(c for c in raw_norm if c.isdigit())
                decimais = None

            if not inteiros and decimais is None:
                return  # campo vazio, não mexe

            # Remove zeros à esquerda, mantém ao menos "0"
            inteiros = inteiros.lstrip("0") or "0"

            # Formata parte inteira com pontos de milhar
            inteiros_fmt = f"{int(inteiros):,}".replace(",", ".")

            if decimais is not None:
                formatado = f"{inteiros_fmt},{decimais}"
            else:
                formatado = inteiros_fmt

            if entry.get() != formatado:
                entry.delete(0, "end")
                entry.insert(0, formatado)
                entry.icursor("end")
        finally:
            _bloqueado[0] = False

    # Vincula ao KeyRelease para formatar após cada tecla
    entry.bind("<KeyRelease>", _formatar)


def valor_do_campo_moeda(entry: ctk.CTkEntry) -> float:
    """Extrai o float de um campo com máscara monetária (ex: '1.234,56' -> 1234.56)."""
    raw = entry.get().strip()
    # Remove R$, espaços, pontos de milhar; troca vírgula decimal por ponto
    limpo = raw.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
    return float(limpo) if limpo else 0.0


def _estilo_cabecalho_excel(ws, cabecalho: list):
    """Aplica estilo de cabeçalho nas células da primeira linha."""
    fill = PatternFill("solid", fgColor="1F538D")
    fonte = Font(bold=True, color="FFFFFF")
    for col_idx, titulo in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = fill
        cell.font = fonte
        cell.alignment = Alignment(horizontal="center")


# ═══════════════════════════════════════════════════════════════════════════
# WIDGET: COMBOBOX COM BUSCA
# ═══════════════════════════════════════════════════════════════════════════

class ComboBoxBusca(ctk.CTkFrame):
    """
    Campo de texto com lista suspensa que filtra em tempo real.
    O dropdown é um Frame com Listbox posicionado via place() sobre
    a janela raiz, evitando qualquer problema com Toplevel no macOS.
    """

    def __init__(self, master, width: int = 260, values: list = None,
                 command=None, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)

        self._todos: list = list(values or [])
        self._command     = command
        self._aberto      = False
        self._frame_dd    = None   # frame do dropdown (place na root)
        self._lb          = None

        self._var = ctk.StringVar()
        self._entry = ctk.CTkEntry(
            self, width=width, textvariable=self._var,
            placeholder_text="Digite para buscar..."
        )
        self._entry.pack(side="left")

        self._var.trace_add("write", self._ao_digitar)
        self._entry.bind("<FocusIn>",  lambda _e: self.after(50, self._abrir))
        self._entry.bind("<FocusOut>", lambda _e: self.after(200, self._checar_fechar))
        self._entry.bind("<Escape>",   lambda _e: self._fechar())
        self._entry.bind("<Down>",     lambda _e: self._focar_lista())
        self._entry.bind("<Return>",   lambda _e: self._fechar())

    # ── Janela raiz ───────────────────────────────────────────────────

    def _janela_raiz(self):
        return self.winfo_toplevel()

    # ── Abertura / fechamento ─────────────────────────────────────────

    def _abrir(self):
        if self._aberto:
            self._popular(self._filtrar())
            return
        self._aberto = True

        self._entry.update_idletasks()
        root = self._janela_raiz()

        # Posição relativa à janela raiz
        ex = self._entry.winfo_rootx() - root.winfo_rootx()
        ey = self._entry.winfo_rooty() - root.winfo_rooty() + self._entry.winfo_height() + 2
        ew = self._entry.winfo_width()

        # Frame container posicionado sobre a root
        self._frame_dd = tk.Frame(root, bg="#2b2b2b", bd=1, relief="solid",
                                  highlightbackground="#1f538d", highlightthickness=1)

        sb = tk.Scrollbar(self._frame_dd, orient="vertical", bg="#2b2b2b")
        sb.pack(side="right", fill="y")

        self._lb = tk_Listbox(
            self._frame_dd,
            yscrollcommand=sb.set,
            bg="#2b2b2b", fg="white",
            selectbackground="#1f538d", selectforeground="white",
            activestyle="dotbox",
            borderwidth=0, highlightthickness=0,
            font=("Helvetica", 12),
            relief="flat",
            exportselection=False,
            height=6,
        )
        self._lb.pack(side="left", fill="both", expand=True)
        sb.config(command=self._lb.yview)

        self._lb.bind("<ButtonRelease-1>", self._ao_selecionar)
        self._lb.bind("<Return>",          self._ao_selecionar)
        self._lb.bind("<Escape>",          lambda _e: self._fechar())
        self._lb.bind("<FocusOut>",        lambda _e: self.after(200, self._checar_fechar))

        self._popular(self._filtrar())

        # Eleva acima de tudo usando place na root
        self._frame_dd.place(x=ex, y=ey, width=ew, height=180)
        self._frame_dd.lift()

    def _fechar(self):
        if self._frame_dd:
            self._frame_dd.place_forget()
            self._frame_dd.destroy()
            self._frame_dd = None
        self._lb     = None
        self._aberto = False

    def _checar_fechar(self):
        if not self._aberto:
            return
        # Se o foco não está nem no entry nem na listbox, fecha
        foco = self.focus_get()
        entry_widget = self._entry._entry if hasattr(self._entry, "_entry") else self._entry
        if foco not in (self._entry, entry_widget, self._lb):
            self._fechar()

    # ── Lista ─────────────────────────────────────────────────────────

    def _popular(self, itens: list):
        if self._lb is None:
            return
        self._lb.delete(0, "end")
        for item in itens:
            self._lb.insert("end", item)

    def _filtrar(self) -> list:
        termo = self._var.get().strip().lower()
        if not termo:
            return self._todos
        return [i for i in self._todos if termo in i.lower()]

    def _focar_lista(self):
        if self._lb and self._lb.size() > 0:
            self._lb.focus_set()
            self._lb.selection_clear(0, "end")
            self._lb.selection_set(0)
            self._lb.activate(0)

    # ── Seleção ───────────────────────────────────────────────────────

    def _ao_digitar(self, *_):
        if not self._aberto:
            self.after(50, self._abrir)
        else:
            self._popular(self._filtrar())

    def _ao_selecionar(self, _e=None):
        if self._lb is None:
            return
        sel = self._lb.curselection()
        if not sel:
            return
        valor = self._lb.get(sel[0])
        self._set_sem_trace(valor)
        self._fechar()
        if self._command:
            self._command(valor)

    # ── API pública ───────────────────────────────────────────────────

    def get(self) -> str:
        return self._var.get()

    def set(self, valor: str):
        self._set_sem_trace(valor)

    def _set_sem_trace(self, valor: str):
        for info in list(self._var.trace_info()):
            try:
                self._var.trace_remove(info[0], info[1])
            except Exception:
                pass
        self._var.set(valor)
        self._var.trace_add("write", self._ao_digitar)

    def configure(self, **kwargs):
        if "values" in kwargs:
            self._todos = list(kwargs.pop("values"))
        if kwargs:
            super().configure(**kwargs)


# ═══════════════════════════════════════════════════════════════════════════
# COMPONENTE: DATEPICKER (calendário popup)
# ═══════════════════════════════════════════════════════════════════════════

class DatePicker:
    """
    Calendário popup posicionado via place() na janela raiz.
    Não usa Toplevel nem grab_set, evitando bloqueio de inputs.
    """

    MESES = ["Janeiro","Fevereiro","Marco","Abril","Maio","Junho",
             "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

    def __init__(self, entry: ctk.CTkEntry):
        self._entry  = entry
        self._frame  = None   # frame colocado na root via place()
        self._ano    = date.today().year
        self._mes    = date.today().month
        self._selecionado = date.today()

    def _root(self):
        return self._entry.winfo_toplevel()

    def toggle(self, *_):
        """Abre ou fecha o calendário."""
        if self._frame and self._frame.winfo_exists():
            self._fechar()
        else:
            self._abrir()

    def _abrir(self):
        # Lê data atual do campo
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(self._entry.get().strip(), "%Y-%m-%d").date()
            self._ano, self._mes, self._selecionado = d.year, d.month, d
        except Exception:
            pass

        root = self._root()
        root.update_idletasks()

        # Calcula posição abaixo do entry, relativa à root
        ex = self._entry.winfo_rootx() - root.winfo_rootx()
        ey = self._entry.winfo_rooty() - root.winfo_rooty() + self._entry.winfo_height() + 2

        self._frame = tk.Frame(root, bg=self._BG, bd=0,
                               highlightbackground=self._BORDER, highlightthickness=1)
        self._renderizar()

        self._frame.update_idletasks()
        w = self._frame.winfo_reqwidth()
        h = self._frame.winfo_reqheight()

        # Garante que não sai pela direita
        rw = root.winfo_width()
        if ex + w > rw:
            ex = max(0, rw - w - 4)

        self._frame.place(x=ex, y=ey, width=w, height=h)
        self._frame.lift()

        # Fecha ao clicar fora
        root.bind("<Button-1>", self._clicar_fora, add="+")

    def _fechar(self):
        if self._frame and self._frame.winfo_exists():
            self._frame.place_forget()
            self._frame.destroy()
        self._frame = None
        try:
            self._root().unbind("<Button-1>")
        except Exception:
            pass

    def _clicar_fora(self, event):
        if self._frame is None:
            return
        try:
            wx = self._frame.winfo_rootx()
            wy = self._frame.winfo_rooty()
            ww = self._frame.winfo_width()
            wh = self._frame.winfo_height()
            # Se o clique foi fora do frame do calendário, fecha
            if not (wx <= event.x_root <= wx + ww and wy <= event.y_root <= wy + wh):
                self._fechar()
        except Exception:
            self._fechar()

    # ── Paleta de cores ──────────────────────────────────────────────────
    _BG        = "#1e1e2e"   # fundo geral
    _BG_CELL   = "#2a2a3d"   # célula normal
    _BG_NAV    = "#252538"   # barra de navegação
    _BG_SEL    = "#3b82f6"   # dia selecionado (azul vivo)
    _BG_TODAY  = "#1e3a5f"   # hoje (azul escuro sutil)
    _BG_HOVER  = "#3b3b55"   # hover célula normal
    _BG_BTN    = "#2e2e45"   # botão "Hoje"
    _FG        = "#e2e8f0"   # texto normal
    _FG_SEL    = "#ffffff"   # texto selecionado
    _FG_TODAY  = "#93c5fd"   # texto hoje
    _FG_HEAD   = "#94a3b8"   # cabeçalho dias semana
    _FG_BTN    = "#60a5fa"   # texto botão "Hoje"
    _BORDER    = "#3b82f6"   # borda do popup

    def _renderizar(self):
        for w in self._frame.winfo_children():
            w.destroy()

        # Borda externa do popup
        self._frame.configure(bg=self._BG, highlightbackground=self._BORDER,
                               highlightthickness=1)

        # Cabeçalho: mês / ano + navegação
        nav = tk.Frame(self._frame, bg=self._BG_NAV)
        nav.grid(row=0, column=0, columnspan=7, sticky="ew", padx=0, pady=0)

        def _nav_btn(parent, txt, cmd):
            b = tk.Label(parent, text=txt, bg=self._BG_NAV, fg=self._FG,
                         font=("Arial", 12, "bold"), cursor="hand2",
                         padx=10, pady=6)
            b.bind("<Button-1>", lambda _: cmd())
            b.bind("<Enter>",    lambda _: b.config(fg=self._FG_BTN))
            b.bind("<Leave>",    lambda _: b.config(fg=self._FG))
            return b

        _nav_btn(nav, "‹", self._mes_anterior).pack(side="left")
        tk.Label(nav, text=f"{self.MESES[self._mes-1]}  {self._ano}",
                 bg=self._BG_NAV, fg=self._FG,
                 font=("Arial", 10, "bold"), pady=8).pack(side="left", expand=True)
        _nav_btn(nav, "›", self._mes_proximo).pack(side="right")

        # Separador
        tk.Frame(self._frame, bg=self._BORDER, height=1).grid(
            row=1, column=0, columnspan=7, sticky="ew")

        # Cabeçalho dos dias da semana
        for col, d in enumerate(["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]):
            tk.Label(self._frame, text=d, bg=self._BG, fg=self._FG_HEAD,
                     font=("Arial", 8, "bold"), width=4, pady=5).grid(
                row=2, column=col, padx=2, pady=(6, 2))

        # Grade de dias
        import calendar as _cal
        semanas = _cal.monthcalendar(self._ano, self._mes)
        hoje = date.today()

        for li, semana in enumerate(semanas):
            for ci, dia in enumerate(semana):
                if dia == 0:
                    tk.Label(self._frame, text="", bg=self._BG, width=4).grid(
                        row=li + 3, column=ci, padx=2, pady=2)
                    continue
                dc = date(self._ano, self._mes, dia)

                if dc == self._selecionado:
                    bg, fg, hover = self._BG_SEL,   self._FG_SEL,  "#2563eb"
                elif dc == hoje:
                    bg, fg, hover = self._BG_TODAY,  self._FG_TODAY, "#1e4a7a"
                else:
                    bg, fg, hover = self._BG_CELL,   self._FG,      self._BG_HOVER

                lbl = tk.Label(self._frame, text=str(dia), bg=bg, fg=fg,
                               font=("Arial", 9), width=4, pady=5,
                               cursor="hand2", relief="flat")
                lbl.grid(row=li + 3, column=ci, padx=2, pady=2)
                lbl.bind("<Button-1>", lambda _e, d=dc: self._selecionar(d))
                lbl.bind("<Enter>",    lambda _e, b=lbl, h=hover: b.config(bg=h))
                lbl.bind("<Leave>",    lambda _e, b=lbl, c=bg:    b.config(bg=c))

        # Separador inferior
        tk.Frame(self._frame, bg=self._BORDER, height=1).grid(
            row=len(semanas) + 3, column=0, columnspan=7, sticky="ew")

        # Botão "Hoje"
        btn = tk.Label(self._frame, text="Hoje", bg=self._BG_BTN, fg=self._FG_BTN,
                       font=("Arial", 8, "bold"), cursor="hand2", pady=6)
        btn.grid(row=len(semanas) + 4, column=0, columnspan=7, sticky="ew", padx=0, pady=0)
        btn.bind("<Button-1>", lambda _: self._selecionar(date.today()))
        btn.bind("<Enter>",    lambda _: btn.config(bg="#383858"))
        btn.bind("<Leave>",    lambda _: btn.config(bg=self._BG_BTN))

    def _mes_anterior(self):
        self._mes, self._ano = (12, self._ano - 1) if self._mes == 1 else (self._mes - 1, self._ano)
        self._renderizar()
        # Reajusta tamanho do frame após re-render
        self._frame.update_idletasks()
        self._frame.place(width=self._frame.winfo_reqwidth(),
                          height=self._frame.winfo_reqheight())

    def _mes_proximo(self):
        self._mes, self._ano = (1, self._ano + 1) if self._mes == 12 else (self._mes + 1, self._ano)
        self._renderizar()
        self._frame.update_idletasks()
        self._frame.place(width=self._frame.winfo_reqwidth(),
                          height=self._frame.winfo_reqheight())

    def _selecionar(self, d: date):
        self._entry.delete(0, "end")
        self._entry.insert(0, d.strftime("%Y-%m-%d"))
        self._fechar()


class CampoData(ctk.CTkFrame):
    """
    Widget composto: CTkEntry (data) + botão calendário em linha.
    Uso direto substituindo o CTkEntry nos formulários.
    O calendário também abre ao clicar no entry.
    """

    def __init__(self, master, width: int = 160, valor_inicial: str = None, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)

        self._ent = ctk.CTkEntry(self, width=width, placeholder_text="AAAA-MM-DD")
        self._ent.pack(side="left")
        if valor_inicial:
            self._ent.insert(0, valor_inicial)

        self._picker = DatePicker(self._ent)

        ctk.CTkButton(self, text="📅", width=34, height=28,
                      fg_color="#333333", hover_color="#444444",
                      command=self._picker.toggle).pack(side="left", padx=(3, 0))

        # Abre calendário ao clicar no entry também
        self._ent.bind("<Button-1>", lambda _: self._picker.toggle())

    # Proxy dos métodos do Entry para compatibilidade com código existente
    def get(self):           return self._ent.get()
    def delete(self, a, b):  self._ent.delete(a, b)
    def insert(self, i, v):  self._ent.insert(i, v)
    def configure(self, **kw): self._ent.configure(**kw)


# ═══════════════════════════════════════════════════════════════════════════
# CAMADA DE BANCO DE DADOS
# ═══════════════════════════════════════════════════════════════════════════

class Database:
    """Gerencia conexão e operações no banco SQLite."""

    def __init__(self, caminho: str = "banco.db"):
        self.conn = sqlite3.connect(caminho, check_same_thread=False)
        self.conn.execute("PRAGMA foreign_keys = ON")
        self._criar_tabelas()

    # ── Criação das tabelas ──────────────────────────────────────────────

    def _criar_tabelas(self):
        """Cria as tabelas caso ainda não existam."""
        cursor = self.conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS clientes (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                nome         TEXT    NOT NULL,
                email        TEXT,
                telefone     TEXT,
                observacoes  TEXT
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS movimentacoes (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente_id  INTEGER NOT NULL,
                tipo        TEXT    NOT NULL CHECK(tipo IN ('Entrada', 'Saida')),
                metodo      TEXT,
                valor       REAL    NOT NULL,
                descricao   TEXT,
                data        TEXT    NOT NULL,
                FOREIGN KEY (cliente_id) REFERENCES clientes(id) ON DELETE CASCADE
            )
        """)
        # Migração: adiciona coluna metodo se ainda não existir (banco legado)
        try:
            self.conn.execute("ALTER TABLE movimentacoes ADD COLUMN metodo TEXT")
            self.conn.commit()
        except Exception:
            pass

        self.conn.commit()

    # ── CRUD Clientes ────────────────────────────────────────────────────

    def listar_clientes(self) -> list:
        cursor = self.conn.execute(
            "SELECT id, nome, email, telefone, observacoes FROM clientes ORDER BY nome"
        )
        return cursor.fetchall()

    def inserir_cliente(self, nome: str, email: str, telefone: str, observacoes: str):
        self.conn.execute(
            "INSERT INTO clientes (nome, email, telefone, observacoes) VALUES (?, ?, ?, ?)",
            (nome, email, telefone, observacoes),
        )
        self.conn.commit()

    def atualizar_cliente(self, id_: int, nome: str, email: str, telefone: str, observacoes: str):
        self.conn.execute(
            "UPDATE clientes SET nome=?, email=?, telefone=?, observacoes=? WHERE id=?",
            (nome, email, telefone, observacoes, id_),
        )
        self.conn.commit()

    def excluir_cliente(self, id_: int):
        self.conn.execute("DELETE FROM clientes WHERE id=?", (id_,))
        self.conn.commit()

    # ── CRUD Movimentações ───────────────────────────────────────────────

    def inserir_movimentacao(
        self, cliente_id: int, tipo: str, metodo: str, valor: float, descricao: str, data: str
    ):
        self.conn.execute(
            "INSERT INTO movimentacoes (cliente_id, tipo, metodo, valor, descricao, data) VALUES (?,?,?,?,?,?)",
            (cliente_id, tipo, metodo, valor, descricao, data),
        )
        self.conn.commit()

    def listar_movimentacoes(self) -> list:
        cursor = self.conn.execute("""
            SELECT m.id, c.nome, m.tipo, m.metodo, m.valor, m.descricao, m.data
            FROM movimentacoes m
            JOIN clientes c ON c.id = m.cliente_id
            ORDER BY m.data DESC, m.id DESC
        """)
        return cursor.fetchall()

    def excluir_movimentacao(self, id_: int):
        self.conn.execute("DELETE FROM movimentacoes WHERE id=?", (id_,))
        self.conn.commit()

    def extrato_cliente(self, cliente_id: int,
                        data_ini: str = None, data_fim: str = None) -> list:
        sql = """
            SELECT tipo, metodo, valor, descricao, data
            FROM movimentacoes
            WHERE cliente_id = ?
        """
        params = [cliente_id]
        if data_ini:
            sql += " AND data >= ?"
            params.append(data_ini)
        if data_fim:
            sql += " AND data <= ?"
            params.append(data_fim)
        sql += " ORDER BY data ASC, id ASC"
        cursor = self.conn.execute(sql, params)
        return cursor.fetchall()

    def saldo_cliente(self, cliente_id: int,
                      data_ini: str = None, data_fim: str = None) -> float:
        """
        Calcula o saldo realizado do cliente.
        Por padrão aplica data_fim = hoje, ignorando movimentações futuras.
        Passe data_fim explicitamente para outro comportamento.
        """
        hoje = date.today().strftime("%Y-%m-%d")
        sql = """
            SELECT
                SUM(CASE WHEN tipo='Entrada' THEN valor ELSE 0 END) -
                SUM(CASE WHEN tipo='Saida'   THEN valor ELSE 0 END)
            FROM movimentacoes
            WHERE cliente_id = ?
        """
        params = [cliente_id]
        if data_ini:
            sql += " AND data >= ?"
            params.append(data_ini)
        # Limita sempre ao máximo entre data_fim informada e hoje
        corte = data_fim if data_fim else hoje
        sql += " AND data <= ?"
        params.append(corte)
        cursor = self.conn.execute(sql, params)
        resultado = cursor.fetchone()[0]
        return resultado if resultado is not None else 0.0

    def saldo_todos_clientes(self) -> list:
        """
        Retorna lista de (id, nome, entradas, saidas, saldo) para todos os clientes.
        Considera apenas movimentações com data <= hoje (saldo realizado).
        """
        hoje = date.today().strftime("%Y-%m-%d")
        cursor = self.conn.execute("""
            SELECT
                c.id,
                c.nome,
                COALESCE(SUM(CASE WHEN m.tipo='Entrada' AND m.data <= ? THEN m.valor ELSE 0 END), 0) AS entradas,
                COALESCE(SUM(CASE WHEN m.tipo='Saida'   AND m.data <= ? THEN m.valor ELSE 0 END), 0) AS saidas,
                COALESCE(SUM(CASE WHEN m.tipo='Entrada' AND m.data <= ? THEN m.valor ELSE 0 END), 0) -
                COALESCE(SUM(CASE WHEN m.tipo='Saida'   AND m.data <= ? THEN m.valor ELSE 0 END), 0) AS saldo
            FROM clientes c
            LEFT JOIN movimentacoes m ON m.cliente_id = c.id
            GROUP BY c.id, c.nome
            ORDER BY c.nome
        """, (hoje, hoje, hoje, hoje))
        return cursor.fetchall()


# ═══════════════════════════════════════════════════════════════════════════
# ABA: CLIENTES
# ═══════════════════════════════════════════════════════════════════════════

class ClientesTab(ctk.CTkFrame):
    """Aba de cadastro e listagem de clientes."""

    def __init__(self, master, db: Database, **kwargs):
        super().__init__(master, **kwargs)
        self.db = db
        self._id_selecionado = None  # ID do cliente em edição

        self._construir_formulario()
        self._construir_tabela()
        self.atualizar_lista()

    # ── Formulário ───────────────────────────────────────────────────────

    def _construir_formulario(self):
        frame = ctk.CTkFrame(self, corner_radius=10)
        frame.pack(fill="x", padx=16, pady=(16, 8))

        ctk.CTkLabel(frame, text="Cadastro de Cliente", font=ctk.CTkFont(size=15, weight="bold")).grid(
            row=0, column=0, columnspan=4, pady=(10, 6), padx=12, sticky="w"
        )

        # Nome
        ctk.CTkLabel(frame, text="Nome *").grid(row=1, column=0, padx=(12, 4), pady=4, sticky="w")
        self.ent_nome = ctk.CTkEntry(frame, width=220, placeholder_text="Nome completo")
        self.ent_nome.grid(row=1, column=1, padx=4, pady=4, sticky="w")

        # E-mail
        ctk.CTkLabel(frame, text="E-mail").grid(row=1, column=2, padx=(16, 4), pady=4, sticky="w")
        self.ent_email = ctk.CTkEntry(frame, width=220, placeholder_text="email@exemplo.com")
        self.ent_email.grid(row=1, column=3, padx=(4, 12), pady=4, sticky="w")

        # Telefone
        ctk.CTkLabel(frame, text="Telefone").grid(row=2, column=0, padx=(12, 4), pady=4, sticky="w")
        self.ent_tel = ctk.CTkEntry(frame, width=220, placeholder_text="(00) 00000-0000")
        self.ent_tel.grid(row=2, column=1, padx=4, pady=4, sticky="w")

        # Observações
        ctk.CTkLabel(frame, text="Observações").grid(row=2, column=2, padx=(16, 4), pady=4, sticky="w")
        self.ent_obs = ctk.CTkEntry(frame, width=220, placeholder_text="Anotações gerais")
        self.ent_obs.grid(row=2, column=3, padx=(4, 12), pady=4, sticky="w")

        # Botões
        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.grid(row=3, column=0, columnspan=4, pady=(6, 12), padx=12, sticky="w")

        ctk.CTkButton(btn_frame, text="Salvar",  width=90,
                      command=self._salvar).pack(side="left", padx=(0, 6))
        ctk.CTkButton(btn_frame, text="Editar",  width=90,
                      fg_color="#2e7d32", hover_color="#388e3c",
                      command=self._editar).pack(side="left", padx=(0, 6))
        ctk.CTkButton(btn_frame, text="Excluir", width=90,
                      fg_color="#c62828", hover_color="#d32f2f",
                      command=self._excluir).pack(side="left", padx=(0, 6))
        ctk.CTkButton(btn_frame, text="Limpar",  width=90,
                      fg_color="gray40", hover_color="gray50",
                      command=self._limpar_form).pack(side="left", padx=(0, 6))

    # ── Tabela ───────────────────────────────────────────────────────────

    def _construir_tabela(self):
        frame = ctk.CTkFrame(self, corner_radius=10)
        frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        ctk.CTkLabel(frame, text="Clientes cadastrados", font=ctk.CTkFont(size=13, weight="bold")).pack(
            anchor="w", padx=12, pady=(10, 4)
        )

        # Estilo para o Treeview seguir o tema escuro
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview",
                         background="#2b2b2b", foreground="white",
                         fieldbackground="#2b2b2b", rowheight=26,
                         borderwidth=0)
        style.configure("Custom.Treeview.Heading",
                         background="#1f538d", foreground="white",
                         relief="flat")
        style.map("Custom.Treeview", background=[("selected", "#1f538d")])

        colunas = ("ID", "Nome", "E-mail", "Telefone", "Observações")
        self.tree = ttk.Treeview(frame, columns=colunas, show="headings",
                                  style="Custom.Treeview", selectmode="browse")

        larguras = {"ID": 40, "Nome": 200, "E-mail": 200, "Telefone": 130, "Observações": 250}
        for col in colunas:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=larguras[col], anchor="center" if col == "ID" else "w")

        scroll = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)

        self.tree.pack(side="left", fill="both", expand=True, padx=(12, 0), pady=(0, 12))
        scroll.pack(side="right", fill="y", pady=(0, 12), padx=(0, 8))

        # Clique em linha popula o formulário
        self.tree.bind("<<TreeviewSelect>>", self._ao_selecionar)

    # ── Eventos ─────────────────────────────────────────────────────────

    def _ao_selecionar(self, _event):
        sel = self.tree.selection()
        if not sel:
            return
        valores = self.tree.item(sel[0], "values")
        self._id_selecionado = int(valores[0])
        self._preencher_form(valores[1], valores[2], valores[3], valores[4])

    def _preencher_form(self, nome, email, telefone, obs):
        self._limpar_form(manter_id=True)
        self.ent_nome.insert(0, nome)
        self.ent_email.insert(0, email)
        self.ent_tel.insert(0, telefone)
        self.ent_obs.insert(0, obs)

    def _limpar_form(self, manter_id=False):
        if not manter_id:
            self._id_selecionado = None
        for ent in (self.ent_nome, self.ent_email, self.ent_tel, self.ent_obs):
            ent.delete(0, "end")

    def _salvar(self):
        nome = self.ent_nome.get().strip()
        if not nome:
            messagebox.showwarning("Campo obrigatorio", "Informe o nome do cliente.")
            return
        self.db.inserir_cliente(
            nome,
            self.ent_email.get().strip(),
            self.ent_tel.get().strip(),
            self.ent_obs.get().strip(),
        )
        print(f"[INFO] Cliente inserido: {nome}")
        self._limpar_form()
        self.atualizar_lista()

    def _editar(self):
        if self._id_selecionado is None:
            messagebox.showwarning("Selecao necessaria", "Selecione um cliente na tabela para editar.")
            return
        nome = self.ent_nome.get().strip()
        if not nome:
            messagebox.showwarning("Campo obrigatorio", "Informe o nome do cliente.")
            return
        self.db.atualizar_cliente(
            self._id_selecionado,
            nome,
            self.ent_email.get().strip(),
            self.ent_tel.get().strip(),
            self.ent_obs.get().strip(),
        )
        print(f"[INFO] Cliente atualizado: id={self._id_selecionado}")
        self._limpar_form()
        self.atualizar_lista()

    def _excluir(self):
        if self._id_selecionado is None:
            messagebox.showwarning("Selecao necessaria", "Selecione um cliente na tabela para excluir.")
            return
        nome = self.ent_nome.get().strip() or f"id={self._id_selecionado}"
        if not messagebox.askyesno("Confirmar exclusao",
                                   f"Excluir o cliente '{nome}' e todas as suas movimentacoes?"):
            return
        self.db.excluir_cliente(self._id_selecionado)
        print(f"[INFO] Cliente excluido: id={self._id_selecionado}")
        self._limpar_form()
        self.atualizar_lista()

    # ── Atualização ──────────────────────────────────────────────────────

    def atualizar_lista(self):
        """Recarrega a tabela com dados atuais do banco."""
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in self.db.listar_clientes():
            self.tree.insert("", "end", values=row)


# ═══════════════════════════════════════════════════════════════════════════
# ABA: MOVIMENTAÇÕES
# ═══════════════════════════════════════════════════════════════════════════

class MovimentacoesTab(ctk.CTkFrame):
    """Aba de cadastro de movimentações financeiras por cliente."""

    def __init__(self, master, db: Database, **kwargs):
        super().__init__(master, **kwargs)
        self.db = db
        self._id_mov_selecionada = None

        self._construir_formulario()
        self._construir_tabela()
        self.atualizar_clientes()
        self.atualizar_lista()

    # ── Formulário ───────────────────────────────────────────────────────

    def _construir_formulario(self):
        frame = ctk.CTkFrame(self, corner_radius=10)
        frame.pack(fill="x", padx=16, pady=(16, 8))

        # Configura colunas: label | campo | label | campo
        frame.columnconfigure(1, weight=1)
        frame.columnconfigure(3, weight=1)

        ctk.CTkLabel(frame, text="Nova Movimentacao", font=ctk.CTkFont(size=15, weight="bold")).grid(
            row=0, column=0, columnspan=4, pady=(10, 6), padx=12, sticky="w"
        )

        # Linha 1: Cliente | Tipo
        ctk.CTkLabel(frame, text="Cliente *").grid(row=1, column=0, padx=(12, 4), pady=4, sticky="w")
        self.cmb_cliente = ComboBoxBusca(frame, width=240, values=[])
        self.cmb_cliente.grid(row=1, column=1, padx=4, pady=4, sticky="w")

        ctk.CTkLabel(frame, text="Tipo *").grid(row=1, column=2, padx=(16, 4), pady=4, sticky="w")
        self.seg_tipo = ctk.CTkSegmentedButton(frame, values=["Entrada", "Saida"])
        self.seg_tipo.set("Entrada")
        self.seg_tipo.grid(row=1, column=3, padx=(4, 12), pady=4, sticky="w")

        # Linha 2: Data | Método
        ctk.CTkLabel(frame, text="Data *").grid(row=2, column=0, padx=(12, 4), pady=4, sticky="w")
        self.ent_data = CampoData(frame, width=160,
                                  valor_inicial=date.today().strftime("%Y-%m-%d"))
        self.ent_data.grid(row=2, column=1, padx=4, pady=4, sticky="w")

        ctk.CTkLabel(frame, text="Metodo").grid(row=2, column=2, padx=(16, 4), pady=4, sticky="w")
        self.seg_metodo = ctk.CTkSegmentedButton(frame, values=["Dinheiro", "PIX", "Debito", "Credito"])
        self.seg_metodo.set("PIX")
        self.seg_metodo.grid(row=2, column=3, padx=(4, 12), pady=4, sticky="w")

        # Linha 3: Valor | Descrição
        ctk.CTkLabel(frame, text="Valor *").grid(row=3, column=0, padx=(12, 4), pady=4, sticky="w")
        self.ent_valor = ctk.CTkEntry(frame, width=240, placeholder_text="Ex: 1500,00")
        self.ent_valor.grid(row=3, column=1, padx=4, pady=4, sticky="w")
        aplicar_mascara_moeda(self.ent_valor)

        ctk.CTkLabel(frame, text="Descricao").grid(row=3, column=2, padx=(16, 4), pady=4, sticky="w")
        self.ent_desc = ctk.CTkEntry(frame, width=240, placeholder_text="Descricao da movimentacao")
        self.ent_desc.grid(row=3, column=3, padx=(4, 12), pady=4, sticky="w")

        # Linha 4: botões — alinhados à esquerda para sempre ficarem visíveis
        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.grid(row=4, column=0, columnspan=4, pady=(6, 12), padx=12, sticky="w")

        ctk.CTkButton(btn_frame, text="Salvar",        width=90,
                      command=self._salvar).pack(side="left", padx=(0, 6))
        ctk.CTkButton(btn_frame, text="Excluir",       width=90,
                      fg_color="#c62828", hover_color="#d32f2f",
                      command=self._excluir).pack(side="left", padx=(0, 6))
        ctk.CTkButton(btn_frame, text="Limpar",        width=90,
                      fg_color="gray40", hover_color="gray50",
                      command=self._limpar_form).pack(side="left", padx=(0, 6))
        ctk.CTkButton(btn_frame, text="Exportar Excel", width=130,
                      fg_color="#1565c0", hover_color="#1976d2",
                      command=self._exportar_excel).pack(side="left", padx=(0, 6))

    # ── Tabela ───────────────────────────────────────────────────────────

    def _construir_tabela(self):
        frame = ctk.CTkFrame(self, corner_radius=10)
        frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        ctk.CTkLabel(frame, text="Movimentacoes registradas", font=ctk.CTkFont(size=13, weight="bold")).pack(
            anchor="w", padx=12, pady=(10, 4)
        )

        style = ttk.Style()
        style.configure("Mov.Treeview",
                         background="#2b2b2b", foreground="white",
                         fieldbackground="#2b2b2b", rowheight=26,
                         borderwidth=0)
        style.configure("Mov.Treeview.Heading",
                         background="#1f538d", foreground="white", relief="flat")
        style.map("Mov.Treeview", background=[("selected", "#1f538d")])

        colunas = ("ID", "Cliente", "Tipo", "Metodo", "Valor (R$)", "Descricao", "Data")
        self.tree = ttk.Treeview(frame, columns=colunas, show="headings",
                                  style="Mov.Treeview", selectmode="browse")

        larguras = {"ID": 40, "Cliente": 160, "Tipo": 75, "Metodo": 90,
                    "Valor (R$)": 100, "Descricao": 200, "Data": 100}
        for col in colunas:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=larguras[col],
                              anchor="center" if col in ("ID", "Tipo", "Metodo", "Valor (R$)", "Data") else "w")

        scroll = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=(12, 0), pady=(0, 12))
        scroll.pack(side="right", fill="y", pady=(0, 12), padx=(0, 8))

        self.tree.bind("<<TreeviewSelect>>", self._ao_selecionar)

    # ── Eventos ─────────────────────────────────────────────────────────

    def _ao_selecionar(self, _event):
        sel = self.tree.selection()
        if not sel:
            return
        valores = self.tree.item(sel[0], "values")
        self._id_mov_selecionada = int(valores[0])

    def _limpar_form(self):
        self._id_mov_selecionada = None
        self.ent_valor.delete(0, "end")
        self.ent_desc.delete(0, "end")
        self.ent_data.delete(0, "end")
        self.ent_data.insert(0, date.today().strftime("%Y-%m-%d"))
        self.seg_tipo.set("Entrada")
        self.seg_metodo.set("PIX")

    def _salvar(self):
        cliente_txt = self.cmb_cliente.get().strip()
        if not cliente_txt:
            messagebox.showwarning("Campo obrigatorio", "Selecione um cliente.")
            return

        try:
            valor = valor_do_campo_moeda(self.ent_valor)
            if valor <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Valor invalido", "Informe um valor numerico positivo.")
            return

        # Valida data
        data_str = self.ent_data.get().strip()
        try:
            from datetime import datetime as _dt
            _dt.strptime(data_str, "%Y-%m-%d")
        except ValueError:
            messagebox.showwarning("Data invalida", "Informe a data no formato AAAA-MM-DD.\nEx: 2025-01-31")
            return

        cliente_id = self._id_do_combo(cliente_txt)
        if cliente_id is None:
            messagebox.showwarning("Cliente invalido", "Selecione um cliente valido.")
            return

        tipo    = self.seg_tipo.get()
        metodo  = self.seg_metodo.get()
        desc    = self.ent_desc.get().strip()

        self.db.inserir_movimentacao(cliente_id, tipo, metodo, valor, desc, data_str)
        self.atualizar_lista()
        self._limpar_form()
        print(f"[INFO] Movimentacao salva: {tipo} {metodo} R${valor:.2f} data={data_str} cliente_id={cliente_id}")


    def _excluir(self):
        if self._id_mov_selecionada is None:
            messagebox.showwarning("Selecao necessaria", "Selecione uma movimentacao na tabela.")
            return
        if not messagebox.askyesno("Confirmar exclusao", "Excluir a movimentacao selecionada?"):
            return
        self.db.excluir_movimentacao(self._id_mov_selecionada)
        print(f"[INFO] Movimentacao excluida: id={self._id_mov_selecionada}")
        self._id_mov_selecionada = None
        self.atualizar_lista()

    # ── Atualização ──────────────────────────────────────────────────────

    def atualizar_clientes(self):
        """Recarrega a lista de clientes no combo."""
        clientes = self.db.listar_clientes()
        self._mapa_clientes = {f"{row[1]} (id={row[0]})": row[0] for row in clientes}
        opcoes = list(self._mapa_clientes.keys())
        self.cmb_cliente.configure(values=opcoes)
        if opcoes:
            self.cmb_cliente.set(opcoes[0])
        else:
            self.cmb_cliente.set("")

    def atualizar_lista(self):
        """Recarrega a tabela de movimentações."""
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in self.db.listar_movimentacoes():
            id_, nome, tipo, metodo, valor, desc, data = row
            self.tree.insert("", "end", values=(
                id_, nome, tipo, metodo or "-", formatar_moeda(valor), desc or "", data
            ))

    def _exportar_excel(self):
        """Exporta todas as movimentações listadas para um arquivo .xlsx."""
        dados = self.db.listar_movimentacoes()
        if not dados:
            messagebox.showinfo("Sem dados", "Nenhuma movimentacao para exportar.")
            return

        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")],
            title="Salvar movimentacoes como",
            initialfile="movimentacoes.xlsx",
        )
        if not caminho:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimentacoes"

        cabecalho = ["ID", "Cliente", "Tipo", "Metodo", "Valor (R$)", "Descricao", "Data"]
        _estilo_cabecalho_excel(ws, cabecalho)

        for row in dados:
            id_, nome, tipo, metodo, valor, desc, data = row
            ws.append([id_, nome, tipo, metodo or "-", valor, desc or "", data])
            ws.cell(row=ws.max_row, column=5).number_format = '#,##0.00'

        # Ajusta largura das colunas automaticamente
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        wb.save(caminho)
        print(f"[INFO] Movimentacoes exportadas para: {caminho}")
        messagebox.showinfo("Exportacao concluida", f"Arquivo salvo em:\n{caminho}")

    def _id_do_combo(self, texto: str):
        return self._mapa_clientes.get(texto)


# ═══════════════════════════════════════════════════════════════════════════
# ABA: EXTRATO
# ═══════════════════════════════════════════════════════════════════════════

class ExtratoTab(ctk.CTkFrame):
    """Aba de extrato financeiro por cliente."""

    def __init__(self, master, db: Database, **kwargs):
        super().__init__(master, **kwargs)
        self.db = db

        self._construir_seletor()
        self._construir_tabela()
        self.atualizar_clientes()

    # ── Seletor de cliente ───────────────────────────────────────────────

    def _construir_seletor(self):
        frame = ctk.CTkFrame(self, corner_radius=10)
        frame.pack(fill="x", padx=16, pady=(16, 8))

        ctk.CTkLabel(frame, text="Extrato por Cliente", font=ctk.CTkFont(size=15, weight="bold")).grid(
            row=0, column=0, columnspan=6, pady=(10, 6), padx=12, sticky="w"
        )

        # Linha 1: cliente + botoes
        ctk.CTkLabel(frame, text="Cliente:").grid(row=1, column=0, padx=(12, 4), pady=4, sticky="w")
        self.cmb_cliente = ComboBoxBusca(frame, width=280, values=[],
                                          command=self._ao_trocar_cliente)
        self.cmb_cliente.grid(row=1, column=1, padx=4, pady=4, sticky="w")

        ctk.CTkButton(frame, text="Atualizar", width=90, command=self._carregar_extrato).grid(
            row=1, column=2, padx=(12, 4), pady=4, sticky="w"
        )
        ctk.CTkButton(frame, text="Exportar Excel", width=120, fg_color="#1565c0", hover_color="#1976d2",
                      command=self._exportar_excel).grid(
            row=1, column=3, padx=(4, 12), pady=4, sticky="w"
        )

        # Saldo
        self.lbl_saldo = ctk.CTkLabel(frame, text="Saldo: R$ 0,00",
                                       font=ctk.CTkFont(size=15, weight="bold"))
        self.lbl_saldo.grid(row=1, column=4, padx=(24, 12), pady=4, sticky="w")

        # Linha 2: filtro de datas com calendário
        ctk.CTkLabel(frame, text="De:").grid(row=2, column=0, padx=(12, 4), pady=(4, 12), sticky="w")
        self.ent_data_ini = CampoData(frame, width=130)
        self.ent_data_ini.grid(row=2, column=1, padx=4, pady=(4, 12), sticky="w")

        ctk.CTkLabel(frame, text="Ate:").grid(row=2, column=2, padx=(12, 4), pady=(4, 12), sticky="w")
        self.ent_data_fim = CampoData(frame, width=130)
        self.ent_data_fim.grid(row=2, column=3, padx=4, pady=(4, 12), sticky="w")

        ctk.CTkButton(frame, text="Filtrar", width=80, command=self._carregar_extrato).grid(
            row=2, column=4, padx=(8, 4), pady=(4, 12), sticky="w"
        )
        ctk.CTkButton(frame, text="Limpar filtro", width=100, fg_color="gray40", hover_color="gray50",
                      command=self._limpar_filtro).grid(
            row=2, column=5, padx=(4, 12), pady=(4, 12), sticky="w"
        )

    # ── Tabela ───────────────────────────────────────────────────────────

    def _construir_tabela(self):
        frame = ctk.CTkFrame(self, corner_radius=10)
        frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        ctk.CTkLabel(frame, text="Movimentacoes do cliente", font=ctk.CTkFont(size=13, weight="bold")).pack(
            anchor="w", padx=12, pady=(10, 4)
        )

        style = ttk.Style()
        style.configure("Ext.Treeview",
                         background="#2b2b2b", foreground="white",
                         fieldbackground="#2b2b2b", rowheight=26,
                         borderwidth=0)
        style.configure("Ext.Treeview.Heading",
                         background="#1f538d", foreground="white", relief="flat")
        style.map("Ext.Treeview", background=[("selected", "#1f538d")])

        colunas = ("Tipo", "Metodo", "Valor (R$)", "Descricao", "Data")
        self.tree = ttk.Treeview(frame, columns=colunas, show="headings",
                                  style="Ext.Treeview", selectmode="none")

        larguras = {"Tipo": 80, "Metodo": 90, "Valor (R$)": 110, "Descricao": 280, "Data": 110}
        for col in colunas:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=larguras[col],
                              anchor="center" if col in ("Tipo", "Metodo", "Valor (R$)", "Data") else "w")

        scroll = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=(12, 0), pady=(0, 12))
        scroll.pack(side="right", fill="y", pady=(0, 12), padx=(0, 8))

        # Tags de cor por tipo
        self.tree.tag_configure("entrada", foreground="#66bb6a")
        self.tree.tag_configure("saida",   foreground="#ef5350")
        self.tree.tag_configure("futuro",  foreground="#888888")  # movimentacao futura (cinza)

    # ── Eventos / Atualização ────────────────────────────────────────────

    def _ao_trocar_cliente(self, _valor=None):
        self._carregar_extrato()

    def _limpar_filtro(self):
        self.ent_data_ini.delete(0, "end")
        self.ent_data_fim.delete(0, "end")
        self._carregar_extrato()

    def atualizar_clientes(self):
        """Recarrega a lista de clientes no combo."""
        clientes = self.db.listar_clientes()
        self._mapa_clientes = {f"{row[1]} (id={row[0]})": row[0] for row in clientes}
        opcoes = list(self._mapa_clientes.keys())
        self.cmb_cliente.configure(values=opcoes)
        if opcoes:
            self.cmb_cliente.set(opcoes[0])
            self._carregar_extrato()
        else:
            self.cmb_cliente.set("")
            self._limpar_tabela()
            self._exibir_saldo(0.0)

    def _carregar_extrato(self):
        """Carrega movimentações do cliente selecionado com filtro de datas opcional."""
        texto = self.cmb_cliente.get().strip()
        cliente_id = self._mapa_clientes.get(texto)
        if cliente_id is None:
            return

        data_ini = self.ent_data_ini.get().strip() or None
        data_fim = self.ent_data_fim.get().strip() or None
        hoje = date.today().strftime("%Y-%m-%d")

        self._limpar_tabela()
        for row in self.db.extrato_cliente(cliente_id, data_ini, data_fim):
            tipo, metodo, valor, descricao, data = row
            futuro = data > hoje
            if futuro:
                tag = "futuro"
                sinal = "+" if tipo == "Entrada" else "-"
                desc_extra = f"{descricao or ''} [agendado]".strip()
            else:
                tag  = "entrada" if tipo == "Entrada" else "saida"
                sinal = "+" if tipo == "Entrada" else "-"
                desc_extra = descricao or ""
            self.tree.insert("", "end", values=(
                tipo, metodo or "-",
                f"{sinal} {formatar_moeda(valor)}",
                desc_extra, data
            ), tags=(tag,))

        saldo = self.db.saldo_cliente(cliente_id, data_ini, data_fim)
        self._exibir_saldo(saldo)

    def _limpar_tabela(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def _exibir_saldo(self, saldo: float):
        cor = "#66bb6a" if saldo >= 0 else "#ef5350"
        self.lbl_saldo.configure(text=f"Saldo: {formatar_moeda(saldo)}", text_color=cor)

    def _exportar_excel(self):
        """Exporta o extrato do cliente selecionado para .xlsx."""
        texto = self.cmb_cliente.get().strip()
        cliente_id = self._mapa_clientes.get(texto)
        if cliente_id is None:
            messagebox.showwarning("Selecao necessaria", "Selecione um cliente para exportar.")
            return

        data_ini = self.ent_data_ini.get().strip() or None
        data_fim = self.ent_data_fim.get().strip() or None

        dados = self.db.extrato_cliente(cliente_id, data_ini, data_fim)
        if not dados:
            messagebox.showinfo("Sem dados", "Este cliente nao possui movimentacoes no periodo.")
            return

        nome_cliente = texto.split(" (id=")[0]
        nome_arquivo = f"extrato_{nome_cliente.replace(' ', '_')}.xlsx"

        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")],
            title="Salvar extrato como",
            initialfile=nome_arquivo,
        )
        if not caminho:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extrato"

        saldo = self.db.saldo_cliente(cliente_id, data_ini, data_fim)
        ws["A1"] = f"Extrato - {nome_cliente}"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = f"Saldo: {formatar_moeda(saldo)}"
        ws["A2"].font = Font(bold=True, color="1B5E20" if saldo >= 0 else "B71C1C")
        ws.append([])  # linha em branco

        cabecalho = ["Tipo", "Valor (R$)", "Descricao", "Data"]
        _estilo_cabecalho_excel(ws, cabecalho)

        cabecalho_ext = ["Tipo", "Metodo", "Valor (R$)", "Descricao", "Data"]
        _estilo_cabecalho_excel(ws, cabecalho_ext)

        fill_entrada = PatternFill("solid", fgColor="E8F5E9")
        fill_saida   = PatternFill("solid", fgColor="FFEBEE")

        for tipo, metodo, valor, descricao, data in dados:
            ws.append([tipo, metodo or "-", valor, descricao or "", data])
            linha = ws.max_row
            ws.cell(linha, 3).number_format = '#,##0.00'
            fill = fill_entrada if tipo == "Entrada" else fill_saida
            for col in range(1, 6):
                ws.cell(linha, col).fill = fill

        # Ajusta largura
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        wb.save(caminho)
        print(f"[INFO] Extrato exportado para: {caminho}")
        messagebox.showinfo("Exportacao concluida", f"Arquivo salvo em:\n{caminho}")


# ═══════════════════════════════════════════════════════════════════════════
# ABA: BALANÇO GERAL
# ═══════════════════════════════════════════════════════════════════════════

class BalancoTab(ctk.CTkFrame):
    """Aba de balanço geral: saldo consolidado de todos os clientes."""

    def __init__(self, master, db: Database, **kwargs):
        super().__init__(master, **kwargs)
        self.db = db
        self._construir_cabecalho()
        self._construir_tabela()
        self.atualizar()

    # ── Cabeçalho ────────────────────────────────────────────────────────

    def _construir_cabecalho(self):
        frame = ctk.CTkFrame(self, corner_radius=10)
        frame.pack(fill="x", padx=16, pady=(16, 8))

        ctk.CTkLabel(frame, text="Balanco Geral",
                     font=ctk.CTkFont(size=15, weight="bold")).grid(
            row=0, column=0, columnspan=4, pady=(10, 6), padx=12, sticky="w")

        ctk.CTkButton(frame, text="Atualizar", width=90,
                      command=self.atualizar).grid(
            row=1, column=0, padx=(12, 4), pady=(0, 12), sticky="w")

        ctk.CTkButton(frame, text="Exportar Excel", width=130,
                      fg_color="#1565c0", hover_color="#1976d2",
                      command=self._exportar_excel).grid(
            row=1, column=1, padx=4, pady=(0, 12), sticky="w")

        # Cards de resumo
        self.lbl_total_entradas = ctk.CTkLabel(
            frame, text="Entradas: R$ 0,00",
            font=ctk.CTkFont(size=13, weight="bold"), text_color="#66bb6a")
        self.lbl_total_entradas.grid(row=1, column=2, padx=(24, 16), pady=(0, 12), sticky="w")

        self.lbl_total_saidas = ctk.CTkLabel(
            frame, text="Saidas: R$ 0,00",
            font=ctk.CTkFont(size=13, weight="bold"), text_color="#ef5350")
        self.lbl_total_saidas.grid(row=1, column=3, padx=(4, 16), pady=(0, 12), sticky="w")

        self.lbl_saldo_geral = ctk.CTkLabel(
            frame, text="Saldo Geral: R$ 0,00",
            font=ctk.CTkFont(size=15, weight="bold"))
        self.lbl_saldo_geral.grid(row=1, column=4, padx=(4, 16), pady=(0, 12), sticky="w")

    # ── Tabela ───────────────────────────────────────────────────────────

    def _construir_tabela(self):
        frame = ctk.CTkFrame(self, corner_radius=10)
        frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        ctk.CTkLabel(frame, text="Saldo por cliente",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(
            anchor="w", padx=12, pady=(10, 4))

        style = ttk.Style()
        style.configure("Bal.Treeview",
                         background="#2b2b2b", foreground="white",
                         fieldbackground="#2b2b2b", rowheight=28, borderwidth=0)
        style.configure("Bal.Treeview.Heading",
                         background="#1f538d", foreground="white", relief="flat")
        style.map("Bal.Treeview", background=[("selected", "#1f538d")])

        colunas = ("Cliente", "Entradas (R$)", "Saidas (R$)", "Saldo (R$)")
        self.tree = ttk.Treeview(frame, columns=colunas, show="headings",
                                  style="Bal.Treeview", selectmode="browse")

        larguras = {"Cliente": 240, "Entradas (R$)": 150, "Saidas (R$)": 150, "Saldo (R$)": 150}
        for col in colunas:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=larguras[col],
                              anchor="w" if col == "Cliente" else "center")

        scroll = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=(12, 0), pady=(0, 12))
        scroll.pack(side="right", fill="y", pady=(0, 12), padx=(0, 8))

        self.tree.tag_configure("positivo", foreground="#66bb6a")
        self.tree.tag_configure("negativo", foreground="#ef5350")
        self.tree.tag_configure("neutro",   foreground="#aaaaaa")

    # ── Atualização ──────────────────────────────────────────────────────

    def atualizar(self):
        """Recarrega os dados do balanço."""
        for item in self.tree.get_children():
            self.tree.delete(item)

        dados = self.db.saldo_todos_clientes()
        total_ent = total_sai = total_sal = 0.0

        for _, nome, entradas, saidas, saldo in dados:
            total_ent += entradas
            total_sai += saidas
            total_sal += saldo
            tag = "positivo" if saldo > 0 else ("negativo" if saldo < 0 else "neutro")
            self.tree.insert("", "end", values=(
                nome,
                formatar_moeda(entradas),
                formatar_moeda(saidas),
                formatar_moeda(saldo),
            ), tags=(tag,))

        self.lbl_total_entradas.configure(text=f"Entradas: {formatar_moeda(total_ent)}")
        self.lbl_total_saidas.configure(text=f"Saidas: {formatar_moeda(total_sai)}")
        cor_saldo = "#66bb6a" if total_sal >= 0 else "#ef5350"
        self.lbl_saldo_geral.configure(
            text=f"Saldo Geral: {formatar_moeda(total_sal)}", text_color=cor_saldo)

    # ── Exportação ───────────────────────────────────────────────────────

    def _exportar_excel(self):
        dados = self.db.saldo_todos_clientes()
        if not dados:
            messagebox.showinfo("Sem dados", "Nenhum dado para exportar.")
            return

        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")],
            title="Salvar balanco como",
            initialfile="balanco_geral.xlsx",
        )
        if not caminho:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balanco Geral"

        # Totais no topo
        total_ent = sum(r[2] for r in dados)
        total_sai = sum(r[3] for r in dados)
        total_sal = sum(r[4] for r in dados)

        ws["A1"] = "Balanco Geral"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Total Entradas: {formatar_moeda(total_ent)}"
        ws["A2"].font = Font(bold=True, color="1B5E20")
        ws["A3"] = f"Total Saidas: {formatar_moeda(total_sai)}"
        ws["A3"].font = Font(bold=True, color="B71C1C")
        ws["A4"] = f"Saldo Geral: {formatar_moeda(total_sal)}"
        ws["A4"].font = Font(bold=True, color="1B5E20" if total_sal >= 0 else "B71C1C")
        ws.append([])

        cabecalho = ["Cliente", "Entradas (R$)", "Saidas (R$)", "Saldo (R$)"]
        _estilo_cabecalho_excel(ws, cabecalho)

        fill_pos = PatternFill("solid", fgColor="E8F5E9")
        fill_neg = PatternFill("solid", fgColor="FFEBEE")
        fill_neu = PatternFill("solid", fgColor="F5F5F5")

        for _, nome, entradas, saidas, saldo in dados:
            ws.append([nome, entradas, saidas, saldo])
            linha = ws.max_row
            for c in range(2, 5):
                ws.cell(linha, c).number_format = '#,##0.00'
            fill = fill_pos if saldo > 0 else (fill_neg if saldo < 0 else fill_neu)
            for c in range(1, 5):
                ws.cell(linha, c).fill = fill

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        wb.save(caminho)
        print(f"[INFO] Balanco exportado para: {caminho}")
        messagebox.showinfo("Exportacao concluida", f"Arquivo salvo em:\n{caminho}")


# ═══════════════════════════════════════════════════════════════════════════
# APLICAÇÃO PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════

class App(ctk.CTk):
    """Janela principal da aplicacao."""

    def __init__(self):
        super().__init__()
        self.title("Client Balance - Controle Financeiro")
        self.geometry("1000x660")
        self.minsize(960, 580)

        # Banco de dados
        self.db = Database("banco.db")
        print("[INFO] Banco de dados conectado: banco.db")

        # Abas
        self.tabview = ctk.CTkTabview(self)
        self.tabview.pack(fill="both", expand=True, padx=10, pady=10)

        self.tabview.add("Clientes")
        self.tabview.add("Movimentacoes")
        self.tabview.add("Extrato")
        self.tabview.add("Balanco")

        self.tab_clientes = ClientesTab(
            self.tabview.tab("Clientes"), db=self.db
        )
        self.tab_clientes.pack(fill="both", expand=True)

        self.tab_mov = MovimentacoesTab(
            self.tabview.tab("Movimentacoes"), db=self.db
        )
        self.tab_mov.pack(fill="both", expand=True)

        self.tab_extrato = ExtratoTab(
            self.tabview.tab("Extrato"), db=self.db
        )
        self.tab_extrato.pack(fill="both", expand=True)

        self.tab_balanco = BalancoTab(
            self.tabview.tab("Balanco"), db=self.db
        )
        self.tab_balanco.pack(fill="both", expand=True)

        # Sincroniza combos ao trocar de aba
        self.tabview.configure(command=self._ao_trocar_aba)

        print("[INFO] Interface carregada com sucesso")

    def _ao_trocar_aba(self):
        """Atualiza os combos de clientes ao navegar pelas abas."""
        aba = self.tabview.get()
        if aba == "Movimentacoes":
            self.tab_mov.atualizar_clientes()
            self.tab_mov.atualizar_lista()
        elif aba == "Extrato":
            self.tab_extrato.atualizar_clientes()
        elif aba == "Balanco":
            self.tab_balanco.atualizar()


# ═══════════════════════════════════════════════════════════════════════════
# PONTO DE ENTRADA
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = App()
    app.mainloop()

